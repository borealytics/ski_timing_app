"""
Calcul des résultats finaux
"""
from typing import List, Dict
from models import Race, Athlete, RunResult
from utils import calculate_best_times, format_time_msscc


class ResultsCalculator:
    """Calcule les résultats finaux d'une course"""
    
    def __init__(self, race: Race):
        self.race = race
    
    def calculate_final_results(self) -> List[Dict]:
        """
        Calcule les résultats finaux selon la méthode de calcul configurée
        
        Returns:
            Liste de résultats triés [{bib, times, total, rank, status}]
        """
        results = []
        
        for athlete in self.race.athletes:
            result = self._calculate_athlete_result(athlete)
            results.append(result)
        
        # Trier les résultats
        results = self._rank_results(results)
        
        return results
    
    def _calculate_athlete_result(self, athlete: Athlete) -> Dict:
        """Calcule le résultat pour un athlète"""
        bib = athlete.bib
        
        # Récupérer les temps de tous les runs
        times = []
        time_displays = []
        statuses = []
        
        for run in self.race.runs:
            result = run.get_result(bib)
            if result:
                times.append(result.time_seconds)
                time_displays.append(result.time_display)
                statuses.append(result.status)
            else:
                times.append(None)
                time_displays.append('')
                statuses.append('PENDING')
        
        # Calculer le temps total selon la méthode
        total_time, status = self._calculate_total(times, statuses)
        
        result_dict = {
            'bib': bib,
            'category': athlete.category,
            'sex': athlete.sex,
            'total_seconds': total_time,
            'total_display': format_time_msscc(total_time) if total_time else status,
            'status': status
        }
        
        # Ajouter les temps individuels
        for i, (time, display) in enumerate(zip(times, time_displays), 1):
            result_dict[f'run{i}'] = display if display else 'PENDING'
            result_dict[f'run{i}_seconds'] = time
        
        return result_dict
    
    def _calculate_total(self, times: List, statuses: List) -> tuple:
        """
        Calcule le temps total selon la méthode configurée

        DSQ, DNF et DNS sont traités de la même façon: le temps de ce run n'est pas valide.
        Le coureur peut quand même être classé s'il a assez de temps valides dans les autres runs.

        Returns:
            (total_seconds, status)
        """
        method = self.race.config.calculation_method

        # Vérifier si tous les runs sont sans temps valide
        if all(s in ('DNS', 'DNF', 'DSQ', 'PENDING') for s in statuses):
            # Déterminer le statut à afficher
            if all(s == 'DNS' for s in statuses):
                return None, 'DNS'
            elif all(s == 'DSQ' for s in statuses):
                return None, 'DSQ'
            elif 'DNF' in statuses:
                return None, 'DNF'
            elif 'DSQ' in statuses:
                return None, 'DSQ'
            else:
                return None, 'DNS'

        # Filtrer les temps valides (ignorer DNS, DNF, DSQ)
        valid_times = [t for t in times if t is not None]

        if method == 'BEST_1':
            if len(valid_times) < 1:
                return None, self._get_failure_status(statuses)
            return min(valid_times), 'FINISHED'

        elif method == 'BEST_2':
            if len(valid_times) < 2:
                return None, self._get_failure_status(statuses)
            sorted_times = sorted(valid_times)
            return sum(sorted_times[:2]), 'FINISHED'

        elif method == 'SUM_3':
            if len(valid_times) < 3:
                return None, self._get_failure_status(statuses)
            return sum(valid_times), 'FINISHED'

        return None, 'UNKNOWN'

    def _get_failure_status(self, statuses: List) -> str:
        """Détermine le statut à afficher quand il n'y a pas assez de temps valides"""
        if 'DNF' in statuses:
            return 'DNF'
        elif 'DSQ' in statuses:
            return 'DSQ'
        elif 'DNS' in statuses:
            return 'DNS'
        return 'DNF'
    
    def _rank_results(self, results: List[Dict]) -> List[Dict]:
        """
        Attribue les rangs aux résultats
        Trie par: Catégorie -> Sexe -> Temps
        """
        # Séparer les résultats valides des invalides
        valid_results = [r for r in results if r['status'] == 'FINISHED']
        invalid_results = [r for r in results if r['status'] != 'FINISHED']
        
        # Trier les résultats valides par catégorie, sexe, puis temps
        valid_results.sort(key=lambda x: (
            x['category'],
            x['sex'],
            x['total_seconds']
        ))
        
        # Attribuer les rangs par catégorie-sexe
        current_category = None
        current_sex = None
        rank = 0
        
        for result in valid_results:
            if (result['category'] != current_category or 
                result['sex'] != current_sex):
                current_category = result['category']
                current_sex = result['sex']
                rank = 1
            else:
                rank += 1
            
            result['rank'] = rank
        
        # Les résultats invalides n'ont pas de rang
        for result in invalid_results:
            result['rank'] = None
        
        # Retourner tous les résultats combinés
        return valid_results + invalid_results
    
    def get_podium_by_category(self, category: str, sex: str, top_n: int = 5) -> List[Dict]:
        """
        Retourne le podium pour une catégorie/sexe donnée
        
        Args:
            category: Catégorie (U6, U8, etc.)
            sex: Sexe (M, F)
            top_n: Nombre de résultats à retourner
            
        Returns:
            Liste des top_n résultats
        """
        all_results = self.calculate_final_results()
        
        # Filtrer par catégorie et sexe
        filtered = [
            r for r in all_results 
            if r['category'] == category and r['sex'] == sex and r['status'] == 'FINISHED'
        ]
        
        return filtered[:top_n]
    
    def export_podiums_to_excel(self, filepath: str):
        """
        Exporte tous les podiums dans un seul fichier Excel, un seul onglet

        Args:
            filepath: Chemin du fichier (avec ou sans extension .xlsx)
        """
        import re
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from pathlib import Path

        all_results = self.calculate_final_results()

        # Obtenir toutes les combinaisons catégorie-sexe triées
        categories = sorted(set(r['category'] for r in all_results),
                          key=lambda c: int(re.search(r'\d+', c).group()) if re.search(r'\d+', c) else 999)
        sexes = sorted(set(r['sex'] for r in all_results))

        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Podiums"

        # Styles
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True, size=11)
        category_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # Titre de la course
        ws.cell(row=row, column=1, value=self.race.config.race_name)
        ws.cell(row=row, column=1).font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2

        # Colonnes
        columns = ['Rang', 'Bib', 'Prénom', 'Nom', 'Club', 'Temps']

        for category in categories:
            for sex in sexes:
                podium = self.get_podium_by_category(category, sex, top_n=5)

                if not podium:
                    continue

                # Titre de la catégorie
                sex_label = "Filles" if sex == "F" else "Garçons"
                ws.cell(row=row, column=1, value=f"{category} - {sex_label}")
                ws.cell(row=row, column=1).font = category_font
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                row += 1

                # En-têtes
                for col_idx, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=row, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                row += 1

                # Données du podium
                for result in podium:
                    athlete = next(a for a in self.race.athletes if a.bib == result['bib'])
                    values = [
                        result['rank'],
                        result['bib'],
                        athlete.first_name,
                        athlete.last_name,
                        athlete.team,
                        result['total_display']
                    ]
                    for col_idx, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.border = thin_border
                        if col_idx in [1, 2, 6]:  # Rang, Bib, Temps centrés
                            cell.alignment = Alignment(horizontal='center')
                    row += 1

                # Ligne vide entre les catégories
                row += 1

        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 8   # Rang
        ws.column_dimensions['B'].width = 8   # Bib
        ws.column_dimensions['C'].width = 15  # Prénom
        ws.column_dimensions['D'].width = 15  # Nom
        ws.column_dimensions['E'].width = 20  # Club
        ws.column_dimensions['F'].width = 12  # Temps

        # S'assurer que le fichier a l'extension .xlsx
        filepath = str(filepath)
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'

        # Sauvegarder
        wb.save(filepath)
        print(f"Exporté: {filepath}")

    def export_full_results_to_excel(self, filepath: str):
        """
        Exporte tous les résultats complets dans un fichier Excel

        Args:
            filepath: Chemin du fichier (avec ou sans extension .xlsx)
        """
        import re
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        all_results = self.calculate_final_results()

        # Obtenir toutes les combinaisons catégorie-sexe triées
        categories = sorted(set(r['category'] for r in all_results),
                          key=lambda c: int(re.search(r'\d+', c).group()) if re.search(r'\d+', c) else 999)
        sexes = sorted(set(r['sex'] for r in all_results))

        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats complets"

        # Styles
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True, size=11)
        category_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        dnf_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # Titre de la course
        ws.cell(row=row, column=1, value=self.race.config.race_name)
        ws.cell(row=row, column=1).font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 2

        # Colonnes de base
        base_columns = ['Rang', 'Bib', 'Prénom', 'Nom', 'Club']

        # Ajouter les colonnes de runs selon le nombre configuré
        run_columns = []
        for i in range(1, self.race.config.num_runs + 1):
            run_columns.append(f'Run {i}')

        columns = base_columns + run_columns + ['Total', 'Status']

        for category in categories:
            for sex in sexes:
                # Filtrer les résultats pour cette catégorie/sexe
                cat_results = [r for r in all_results
                              if r['category'] == category and r['sex'] == sex]

                if not cat_results:
                    continue

                # Titre de la catégorie
                sex_label = "Filles" if sex == "F" else "Garçons"
                ws.cell(row=row, column=1, value=f"{category} - {sex_label}")
                ws.cell(row=row, column=1).font = category_font
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
                row += 1

                # En-têtes
                for col_idx, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=row, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                row += 1

                # Données
                for result in cat_results:
                    athlete = next(a for a in self.race.athletes if a.bib == result['bib'])

                    values = [
                        result['rank'] if result['rank'] else '',
                        result['bib'],
                        athlete.first_name,
                        athlete.last_name,
                        athlete.team
                    ]

                    # Ajouter les temps de chaque run
                    for i in range(1, self.race.config.num_runs + 1):
                        run_key = f'run{i}'
                        values.append(result.get(run_key, ''))

                    values.append(result['total_display'])
                    values.append(result['status'])

                    is_dnf = result['status'] in ('DNF', 'DNS', 'DSQ')

                    for col_idx, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.border = thin_border
                        if col_idx in [1, 2] or col_idx > len(base_columns):
                            cell.alignment = Alignment(horizontal='center')
                        if is_dnf:
                            cell.fill = dnf_fill
                    row += 1

                # Ligne vide entre les catégories
                row += 1

        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 8   # Rang
        ws.column_dimensions['B'].width = 8   # Bib
        ws.column_dimensions['C'].width = 15  # Prénom
        ws.column_dimensions['D'].width = 15  # Nom
        ws.column_dimensions['E'].width = 20  # Club

        # Colonnes de runs et total
        col_letters = ['F', 'G', 'H', 'I', 'J', 'K']
        for i, letter in enumerate(col_letters[:self.race.config.num_runs + 2]):
            ws.column_dimensions[letter].width = 12

        # S'assurer que le fichier a l'extension .xlsx
        filepath = str(filepath)
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'

        # Sauvegarder
        wb.save(filepath)
        print(f"Exporté: {filepath}")
